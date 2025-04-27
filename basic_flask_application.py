import json
from flask import Flask, request, jsonify
import os
import openpyxl


# --- Load Data ---
def load_leaderboard(filename="leaderboard.xlsx"):
   try:
       workbook = openpyxl.load_workbook(filename)
       sheet = workbook.active
       leaderboard = {}
       for row in sheet.iter_rows(min_row=2, values_only=True):
           name, score = row
           leaderboard[name] = score
       return leaderboard
   except FileNotFoundError:
       return {}

def load_teams(filename="teams.xlsx"):
   try:
       workbook = openpyxl.load_workbook(filename)
       sheet = workbook.active
       teams = {}
       for row in sheet.iter_rows(min_row=2, values_only=True):
           team, score, members = row
           teams[team] = {
               'score': score,
               'members': set(m.strip() for m in members.split(',')) if members else set()
           }
       return teams
   except FileNotFoundError:
       return {}

DATA_FILE = 'scores.json'
# Load existing scores from JSON file or create a new one
if os.path.exists(DATA_FILE):
    with open(DATA_FILE, 'r') as f:
        scores = json.load(f)
else:
    scores = {}

leaderboard = load_leaderboard()
teams = load_teams()
app = Flask(__name__)

@app.route('/')
def home():
    return jsonify(scores)

@app.route('/update', methods=['POST'])
def update_score():
    data = request.get_json()
    name = data.get('name')
    score = data.get('score', 0)
    team = data.get('team')

    if name:
        if name in scores:
            scores[name] += score
        else:
            scores[name] = score

    if team not in teams:
        teams[team] = {'members': set(), 'score': 0}

    if len(teams[team]['members']) < 4 or name in teams[team]['members']:
        teams[team]['members'].add(name)
        teams[team]['score'] += score

        with open(DATA_FILE, 'w') as f:
            json.dump(scores, f)

        return jsonify({'message': f"Added {score} points to {name}. New total: {scores[name]}"}), 200
    else:
        return jsonify({'error': 'No name provided'}), 400

@app.route('/leaderboard')
def show_leaderboards():
    sorted_individuals = dict(sorted(scores.items(), key=lambda item: item[1], reverse=True))
    sorted_teams = dict(sorted(teams.items(), key=lambda item: item[1]['score'], reverse=True))

    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Leaderboards</title>
        <style>
            body { font-family: Arial, sans-serif; padding: 20px; background: #f0f0f0; }
            h1 { text-align: center; }
            .container { display: flex; justify-content: space-around; gap: 40px; flex-wrap: wrap; }
            table {
                background: white; border-collapse: collapse; width: 100%; max-width: 400px;
                box-shadow: 0 2px 5px rgba(0,0,0,0.1);
            }
            th, td { padding: 10px; border-bottom: 1px solid #ddd; text-align: left; }
            th { background-color: #4CAF50; color: white; }
        </style>
    </head>
    <body>
        <h1>Game Leaderboards</h1>
        <div class="container">
            <div>
                <h2>Individual Leaderboard</h2>
                <table>
                    <tr><th>Player</th><th>Score</th></tr>
    """

    for name, score in sorted_individuals.items():
        html += f"<tr><td>{name}</td><td>{score}</td></tr>"

    html += """
                </table>
            </div>
            <div>
                <h2>Team Leaderboard</h2>
                <table>
                    <tr><th>Team</th><th>Score</th></tr>
    """

    for team, data in sorted_teams.items():
        html += f"<tr><td>{team}</td><td>{data['score']}</td></tr>"

    html += """
                </table>
            </div>
        </div>
    </body>
    </html>
    """
    return html


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
